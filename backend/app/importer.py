"""Excel/CSV import for the Buyer roster.

Accepts a spreadsheet whose header row maps (case-insensitively, fuzzily) to the
Buyer fields. Unknown columns are ignored; missing optional fields default empty.
"""
from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

from .schemas import BuyerCreate

# Map many possible header spellings -> canonical field name.
_ALIASES = {
    "name": "name", "buyer": "name", "company": "name", "buyer name": "name", "entity": "name",
    "sector": "sector", "industry": "sector",
    "products": "products", "product": "products", "makes": "products", "output": "products",
    "raw materials needed": "raw_materials_needed", "raw materials": "raw_materials_needed",
    "inputs": "raw_materials_needed", "needs": "raw_materials_needed",
    "customers served": "customers_served", "customers": "customers_served", "sells to": "customers_served",
    "acquisition thesis": "acquisition_thesis", "thesis": "acquisition_thesis", "mandate": "acquisition_thesis",
    "surplus cash": "surplus_cash_inr_cr", "surplus cash inr cr": "surplus_cash_inr_cr",
    "cash": "surplus_cash_inr_cr", "budget": "surplus_cash_inr_cr",
    "geography": "geography_pref", "geography pref": "geography_pref", "location": "geography_pref",
    "notes": "notes", "remarks": "notes",
}


def _canon(header: str) -> str | None:
    return _ALIASES.get((header or "").strip().lower())


def _row_to_buyer(headers: list[str], values: list) -> BuyerCreate | None:
    data: dict = {}
    for h, v in zip(headers, values):
        field = _canon(h)
        if not field:
            continue
        val = "" if v is None else str(v).strip()
        if field == "surplus_cash_inr_cr":
            try:
                data[field] = float(str(v).replace(",", "")) if v not in (None, "") else 0.0
            except ValueError:
                data[field] = 0.0
        else:
            data[field] = val
    if not data.get("name"):
        return None
    return BuyerCreate(**data)


def parse_xlsx(content: bytes) -> list[BuyerCreate]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    out: list[BuyerCreate] = []
    for values in rows:
        b = _row_to_buyer(headers, list(values))
        if b:
            out.append(b)
    return out


def parse_csv(content: bytes) -> list[BuyerCreate]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        return []
    out: list[BuyerCreate] = []
    for values in reader:
        b = _row_to_buyer(headers, values)
        if b:
            out.append(b)
    return out


def parse_upload(filename: str, content: bytes) -> list[BuyerCreate]:
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return parse_xlsx(content)
    if filename.lower().endswith(".csv"):
        return parse_csv(content)
    # try xlsx then csv
    try:
        return parse_xlsx(content)
    except Exception:
        return parse_csv(content)
