"""Build the CONFIDENTIAL master.html (run locally, kept private by KCM).

Reads the client-master Excel, encrypts it (AES-256-GCM, key = PBKDF2 of the
master password — matches the Web-Crypto code in master.js), and inlines it,
plus a snapshot of the public IBBI feed, into a single self-contained
master.html. The plaintext client list and the password are NEVER written to
disk; only ciphertext is embedded.

Run:  set MASTER_PASSWORD in backend/.env, then  python -m app.build_master
"""
import base64
import hashlib
import json
import os
from pathlib import Path

from openpyxl import load_workbook

from .config import settings

ROOT = Path(__file__).resolve().parents[2]
FRONTEND = Path(__file__).resolve().parents[1] / "frontend"
SITE_DATA = ROOT / "site" / "data.json"
OUT_DIR = ROOT / "master"                      # gitignored — private deliverable
OUT = OUT_DIR / "master.html"

SHEET = "Working Base - All Clients"
ITERS = 200000

# Excel column header -> client field
COLMAP = {
    "name": "Company",
    "sector": "Industry",
    "products": "Products / Services",
    "customers_served": "End-markets",
    "value_chain": "Value-chain position",
    "synergy_directions": "Synergy directions",
    "eligibility_29a": "29A flags",
    "confidence": "Confidence",
    "reach": "Decision locus",
    "notes": "Business (Layer 2)",
}


def _cell(row, idx, header):
    i = idx.get(header)
    if i is None or i >= len(row) or row[i] is None:
        return ""
    return str(row[i]).strip()


def read_clients(xlsx: Path) -> list[dict]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}
    out = []
    for row in rows:
        name = _cell(row, idx, "Company")
        if not name:
            continue
        c = {field: _cell(row, idx, header) for field, header in COLMAP.items()}
        thesis = " — ".join(x for x in [_cell(row, idx, "Acquirer-fit"), c["synergy_directions"]] if x)
        c["acquisition_thesis"] = thesis[:400]
        out.append(c)
    wb.close()
    return out


def encrypt_obj(obj: dict, password: str, iters: int = ITERS) -> dict:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    salt, iv = os.urandom(16), os.urandom(12)
    key = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iters, dklen=32)
    ct = AESGCM(key).encrypt(iv, json.dumps(obj, ensure_ascii=False).encode("utf-8"), None)
    b = base64.b64encode
    return {"v": 1, "iter": iters, "salt": b(salt).decode(), "iv": b(iv).decode(), "ct": b(ct).decode()}


def ensure_feed() -> dict:
    if not SITE_DATA.exists():
        from .build_site import main as build_public
        build_public()
    return json.loads(SITE_DATA.read_text(encoding="utf-8"))


def main() -> None:
    password = (os.environ.get("MASTER_PASSWORD") or settings.master_password or "").strip()
    if not password:
        raise SystemExit("MASTER_PASSWORD is not set. Put it in backend/.env (gitignored) and retry.")
    xlsx = Path((os.environ.get("MASTER_XLSX") or settings.master_xlsx or "").strip() or
                r"D:\OneDrive - K C Mehta & Co LLP\Mehta, Suril- KCM's files - General but confidential\@Research\IBC Companies Search\KCM_IBC_Client_Master.xlsx")
    if not xlsx.exists():
        raise SystemExit(f"Client-master Excel not found: {xlsx}\nSet MASTER_XLSX to its path.")
    feed_url = (os.environ.get("MASTER_FEED_URL") or settings.master_feed_url or "").strip() or "data.json"

    clients = read_clients(xlsx)
    blob = encrypt_obj({"clients": clients}, password)
    feed = ensure_feed()

    tpl = (FRONTEND / "master_template.html").read_text(encoding="utf-8")
    core_css = (FRONTEND / "core.css").read_text(encoding="utf-8")
    core_js = (FRONTEND / "core.js").read_text(encoding="utf-8")
    master_js = (FRONTEND / "master.js").read_text(encoding="utf-8")

    html = (tpl
            .replace("/*__CORE_CSS__*/", core_css)
            .replace("/*__CORE_JS__*/", core_js)
            .replace("/*__MASTER_JS__*/", master_js)
            .replace("__DATA_URL__", feed_url)
            .replace("__ENGAGE_FORM_URL__", (os.environ.get("ENGAGE_FORM_URL") or settings.engage_form_url or "").strip())
            .replace("__ENGAGE_KEY__", (os.environ.get("ENGAGE_KEY") or settings.engage_key or "").strip())
            .replace("__FEED__", json.dumps(feed, ensure_ascii=False))
            .replace("__BLOB__", json.dumps(blob)))

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT.write_text(html, encoding="utf-8")
    size_mb = OUT.stat().st_size / (1024 * 1024)
    print(f"[master] encrypted {len(clients):,} clients (AES-256-GCM, PBKDF2 {ITERS} iters)")
    print(f"[master] embedded feed: {len(feed.get('opportunities', []))} companies; data_url={feed_url}")
    print(f"[master] wrote {OUT} ({size_mb:.2f} MB) — keep this file PRIVATE.")


if __name__ == "__main__":
    main()
